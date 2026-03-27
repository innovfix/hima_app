from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path


@dataclass(frozen=True)
class TestCase:
    id: str
    title: str
    module: str
    severity: str
    priority: str
    preconditions: str
    steps: str
    test_data: str
    expected_result: str
    actual_result: str = ""
    notes: str = ""


def _today_stamp() -> str:
    return date.today().isoformat()


def build_testcases() -> list[TestCase]:
    return [
        TestCase(
            id="HIMA-SPRINT-001",
            title="Session persists after tab navigation and app relaunch",
            module="Session / Auth",
            severity="High",
            priority="High",
            preconditions="Valid user account. Fresh login.",
            steps=(
                "1. Log in with valid credentials.\n"
                "2. Navigate in sequence: Home → New → Favorite → Profile (repeat 2–3 cycles).\n"
                "3. Remove app from recents / force-stop the app (or swipe away from background).\n"
                "4. Relaunch the app from the launcher."
            ),
            test_data="N/A",
            expected_result=(
                "User remains logged in after relaunch.\n"
                "App opens to Home (or last intended landing) without forcing login again, "
                "unless the user had explicitly logged out."
            ),
            notes="Regression: verify no unexpected logout after navigation + cold start.",
        ),
        TestCase(
            id="HIMA-SPRINT-002",
            title="Explicit logout still requires login on next launch",
            module="Session / Auth",
            severity="High",
            priority="High",
            preconditions="User logged in.",
            steps=(
                "1. From Profile/Settings, tap Logout and confirm.\n"
                "2. Close and relaunch the app."
            ),
            test_data="N/A",
            expected_result="User is not logged in; login screen or flow is shown as designed.",
            notes="Negative check alongside HIMA-SPRINT-001 to avoid false positives.",
        ),
        TestCase(
            id="HIMA-SPRINT-003",
            title="New tab shows populated content after login",
            module="New tab",
            severity="High",
            priority="High",
            preconditions="Log in with valid credentials.",
            steps=(
                "1. After successful login, tap the New tab.\n"
                "2. Wait for loading to complete (pull-to-refresh if available).\n"
                "3. Scroll the list/grid."
            ),
            test_data="Account with expected “new” content on server.",
            expected_result=(
                "New tab is not blank: user sees new profiles or new content per product design "
                "(list/cards/empty-state with message—not a white screen)."
            ),
        ),
        TestCase(
            id="HIMA-SPRINT-004",
            title="Wallet opens without crash after heavy tab switching",
            module="Wallet",
            severity="Critical",
            priority="High",
            preconditions="Logged in. Wallet entry point available (e.g., from Profile or menu).",
            steps=(
                "1. Switch tabs repeatedly: Home, Favorite, Recent (if present), Profile, New (10+ switches).\n"
                "2. Open Wallet / Wallet & Coins from the defined navigation path.\n"
                "3. Observe screen load and interact (scroll if possible)."
            ),
            test_data="N/A",
            expected_result=(
                "Wallet screen opens successfully; no crash, ANR, or black screen.\n"
                "Content loads and remains usable."
            ),
            notes="If crash occurs, capture logcat / stack trace and device model.",
        ),
        TestCase(
            id="HIMA-SPRINT-005",
            title="Rapid taps on same tab do not blank the screen",
            module="Bottom navigation / Tabs",
            severity="High",
            priority="High",
            preconditions="Logged in. On main shell with bottom tabs.",
            steps=(
                "1. Go to Home (or Favorite / Profile / Recent).\n"
                "2. Rapidly tap the same tab icon 10–15 times in quick succession.\n"
                "3. Observe content area after tapping stops.\n"
                "4. Repeat for other main tabs."
            ),
            test_data="N/A",
            expected_result=(
                "Content does not disappear; screen does not become blank.\n"
                "If re-selection reloads data, UI should show loading then content—not empty void."
            ),
        ),
        TestCase(
            id="HIMA-SPRINT-006",
            title="Help & Support: Hindi category shows Hindi-only content",
            module="Help & Support",
            severity="Medium",
            priority="Medium",
            preconditions="Logged in. Help & Support accessible.",
            steps=(
                "1. Open Help & Support.\n"
                "2. Select Hindi as the language/category (per your UI).\n"
                "3. Open multiple articles/sections under Hindi.\n"
                "4. Scan full visible text for mixed languages."
            ),
            test_data="N/A",
            expected_result=(
                "All displayed help content for Hindi selection is in Hindi only; "
                "no English/other-language fragments unless explicitly required by design."
            ),
        ),
        TestCase(
            id="HIMA-SPRINT-007",
            title="Incoming call notification opens correct incoming-call UI",
            module="Calls / Notifications",
            severity="High",
            priority="High",
            preconditions=(
                "Two test accounts/devices. Callee logged in; caller can place a call.\n"
                "Notification permission granted for the app."
            ),
            steps=(
                "1. From caller account, start a call to the logged-in user.\n"
                "2. On callee device, wait for incoming call notification.\n"
                "3. Tap the notification (not Accept from notification action if separate).\n"
                "4. Observe which screen opens."
            ),
            test_data="Two valid users; active call permission.",
            expected_result=(
                "App foregrounds to the incoming call screen for that call (accept/reject visible).\n"
                "User is not stuck on Home or unrelated screen without call UI."
            ),
            notes="Test with app in background and killed if your flow supports it.",
        ),
        TestCase(
            id="HIMA-SPRINT-008",
            title="Toast size is compact and within UI guidelines",
            module="UI / Toasts",
            severity="Low",
            priority="Low",
            preconditions="Any flow that shows a toast (error, info, success).",
            steps=(
                "1. Trigger a toast (e.g., validation error, success message).\n"
                "2. Observe height, width, padding, and text wrapping.\n"
                "3. Compare with design reference if available."
            ),
            test_data="Any action that reliably shows toast.",
            expected_result=(
                "Toast is not oversized; does not dominate the screen; "
                "readable with normal padding and duration."
            ),
        ),
        TestCase(
            id="HIMA-SPRINT-009",
            title="Profile icon tap: no meaningless toast",
            module="Profile",
            severity="Low",
            priority="Low",
            preconditions="Logged in. User Profile icon visible (e.g., toolbar).",
            steps=(
                "1. Tap the User Profile icon once.\n"
                "2. Navigate back and tap again 3–5 times.\n"
                "3. Note any toast shown."
            ),
            test_data="N/A",
            expected_result=(
                "No toast on every tap unless it conveys useful context; "
                "or no toast at all if navigation alone is sufficient."
            ),
        ),
        TestCase(
            id="HIMA-SPRINT-010",
            title="User profile: last card aligns with sections above",
            module="Profile screen",
            severity="Low",
            priority="Low",
            preconditions="Logged in. Open full user profile (own or other per app).",
            steps=(
                "1. Scroll to the bottom of the profile page.\n"
                "2. Compare horizontal alignment and margins of the last card vs cards above.\n"
                "3. Check in portrait; optionally landscape."
            ),
            test_data="Profile with enough sections to show last card.",
            expected_result=(
                "Last card matches width, padding, and vertical spacing of previous cards; "
                "no misaligned edge or odd bottom gap."
            ),
        ),
        TestCase(
            id="HIMA-SPRINT-011",
            title="Wallet / Coins: label and amount have clear spacing",
            module="Wallet > Coins",
            severity="Low",
            priority="Low",
            preconditions="Logged in. Navigate to Wallet → Coins subsection.",
            steps=(
                "1. Locate coin balance label and amount.\n"
                "2. Check spacing between label and value on multiple rows if present.\n"
                "3. Verify readability at default and large font (accessibility)."
            ),
            test_data="Account with visible coin balance.",
            expected_result=(
                "Adequate spacing between label and amount; visually balanced; no cramped overlap."
            ),
        ),
        TestCase(
            id="HIMA-SPRINT-012",
            title="Offline: clear message and Retry when network is lost",
            module="Connectivity",
            severity="High",
            priority="High",
            preconditions="Logged in. App in foreground.",
            steps=(
                "1. With Wi‑Fi and mobile data on, open Home.\n"
                "2. Turn off Wi‑Fi and mobile data (airplane mode or toggles).\n"
                "3. Try Home, Profile, Call entry, and one other feature.\n"
                "4. Turn network back on; tap Retry if shown."
            ),
            test_data="N/A",
            expected_result=(
                "App detects offline state.\n"
                "User sees a clear “No Internet” (or equivalent) message.\n"
                "Retry is visible and works: content reloads after connectivity returns."
            ),
        ),
    ]


def write_xlsx(out_path: Path, testcases: list[TestCase], sheet_title: str) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:  # pragma: no cover
        raise SystemExit(
            "Missing dependency: openpyxl. Install: pip install openpyxl "
            "or use: PYTHONPATH=qa/.vendor python3 qa/generate_sprint_bug_testcases_xlsx.py"
        ) from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name limit

    headers = [
        "Test Case ID",
        "Title",
        "Module",
        "Severity",
        "Priority",
        "Preconditions",
        "Steps",
        "Test Data",
        "Expected Result",
        "Actual Result",
        "Notes",
    ]

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    body_alignment = Alignment(vertical="top", wrap_text=True)
    for tc in testcases:
        ws.append(
            [
                tc.id,
                tc.title,
                tc.module,
                tc.severity,
                tc.priority,
                tc.preconditions,
                tc.steps,
                tc.test_data,
                tc.expected_result,
                tc.actual_result,
                tc.notes,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = body_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"

    widths = {
        "A": 16,
        "B": 46,
        "C": 26,
        "D": 12,
        "E": 12,
        "F": 38,
        "G": 44,
        "H": 22,
        "I": 46,
        "J": 22,
        "K": 26,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 26

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    out_file = repo_root / "qa" / "testcases" / f"HIMA_SprintBugs_TestCases_{_today_stamp()}.xlsx"
    testcases = build_testcases()
    write_xlsx(out_file, testcases, "Sprint Bug Test Cases")
    print(str(out_file))


if __name__ == "__main__":
    main()
