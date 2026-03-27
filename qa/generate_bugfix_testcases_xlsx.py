from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path


@dataclass(frozen=True)
class TestCase:
    id: str
    title: str
    module: str
    severity: str
    priority: str
    preconditions: str
    steps: str
    test_data: str
    expected_result: str
    actual_result: str = ""
    notes: str = ""


def _today_stamp() -> str:
    return date.today().isoformat()


def build_testcases() -> list[TestCase]:
    # Each item maps to the bug list provided by the user, plus a few key regressions.
    return [
        # 1) Disabled call/video call buttons still clickable without feedback
        TestCase(
            id="HIMA-BUGFIX-001",
            title="Call button disabled: tap shows feedback or no action",
            module="User Profile",
            severity="Medium",
            priority="Medium",
            preconditions=(
                "Login with a valid user.\n"
                "Navigate to a user profile where Call is disabled (visually disabled state)."
            ),
            steps=(
                "1. Observe Call button is visually disabled.\n"
                "2. Tap the Call button once.\n"
                "3. Tap the Call button multiple times rapidly."
            ),
            test_data="A profile where Call is disabled.",
            expected_result=(
                "Call button should either be non-interactive (no click triggered) OR show a user-friendly message "
                "explaining why calls are unavailable (e.g., “This user is not accepting calls right now”).\n"
                "No call should be initiated."
            ),
            notes="Pass if user gets clear feedback and call does not start.",
        ),
        TestCase(
            id="HIMA-BUGFIX-002",
            title="Video Call button disabled: tap shows feedback or no action",
            module="User Profile",
            severity="Medium",
            priority="Medium",
            preconditions=(
                "Login with a valid user.\n"
                "Navigate to a user profile where Video Call is disabled (visually disabled state)."
            ),
            steps=(
                "1. Observe Video Call button is visually disabled.\n"
                "2. Tap the Video Call button once.\n"
                "3. Tap the Video Call button multiple times rapidly."
            ),
            test_data="A profile where Video Call is disabled.",
            expected_result=(
                "Video Call button should either be non-interactive OR show a user-friendly message explaining why "
                "video calls are unavailable.\n"
                "No video call should be initiated."
            ),
            notes="Pass if user gets clear feedback and video call does not start.",
        ),
        TestCase(
            id="HIMA-BUGFIX-003",
            title="Call/Video enabled: buttons initiate call flows normally",
            module="User Profile",
            severity="Medium",
            priority="Medium",
            preconditions=(
                "Login with a valid user.\n"
                "Navigate to a user profile where Call and Video Call are enabled."
            ),
            steps=(
                "1. Tap Call.\n"
                "2. Return to profile.\n"
                "3. Tap Video Call."
            ),
            test_data="A profile where Call/Video Call are enabled.",
            expected_result="Call and Video Call should work as per normal app behavior (no false-disabled behavior).",
            notes="Regression check to ensure no accidental disable.",
        ),
        # 2) Timer icon misaligned with call duration in Recent Session / Call History
        TestCase(
            id="HIMA-BUGFIX-004",
            title="Recent Calls: timer icon vertically aligned with duration text",
            module="Recent Session / Call History",
            severity="Low",
            priority="Low",
            preconditions="Login with a valid user and open Recent Session / Call History.",
            steps=(
                "1. Locate at least 5 call history entries with duration + timer icon.\n"
                "2. Compare icon alignment relative to the duration baseline and adjacent text.\n"
                "3. Scroll list and verify alignment remains consistent."
            ),
            test_data="Call history entries with visible duration.",
            expected_result="Timer icon is visually aligned with duration text consistently across items and after scrolling.",
        ),
        TestCase(
            id="HIMA-BUGFIX-005",
            title="Recent Calls: timer icon alignment consistent across font scales",
            module="Recent Session / Call History",
            severity="Low",
            priority="Low",
            preconditions="Device accessibility font scale set to Large (or 1.3x+). User logged in.",
            steps=(
                "1. Open Recent Session / Call History.\n"
                "2. Observe timer icon alignment next to duration text for multiple rows."
            ),
            test_data="Android font scale: Large.",
            expected_result="Timer icon remains aligned; no clipping or misplacement at larger font scale.",
            notes="Optional but recommended UI robustness check.",
        ),
        # 3) Unexpected toast on long press bottom navigation tab icons
        TestCase(
            id="HIMA-BUGFIX-006",
            title="Bottom nav: long press shows no toast for Home tab",
            module="Bottom Navigation",
            severity="Low",
            priority="Low",
            preconditions="Login and land on main screen with bottom navigation visible.",
            steps="1. Long-press the Home tab icon for ~1 second.",
            test_data="N/A",
            expected_result="No toast (or tooltip/toast) should appear on long press.",
        ),
        TestCase(
            id="HIMA-BUGFIX-007",
            title="Bottom nav: long press shows no toast for Favorite/New/Profile tabs",
            module="Bottom Navigation",
            severity="Low",
            priority="Low",
            preconditions="Login and land on main screen with bottom navigation visible.",
            steps=(
                "1. Long-press the Favorite tab icon for ~1 second.\n"
                "2. Long-press the New tab icon for ~1 second.\n"
                "3. Long-press the Profile tab icon for ~1 second."
            ),
            test_data="N/A",
            expected_result="No toast (or tooltip/toast) should appear on long press for any tab.",
        ),
        # 4) "How to Get Coins" option misaligned in Refer & Earn
        TestCase(
            id="HIMA-BUGFIX-008",
            title="Refer & Earn: 'How to Get Coins' option aligned per design",
            module="Refer & Earn",
            severity="Low",
            priority="Low",
            preconditions="Login with a valid user and navigate to Refer & Earn screen.",
            steps=(
                "1. Locate 'How to Get Coins' option.\n"
                "2. Observe horizontal alignment with other items and overall spacing.\n"
                "3. Rotate device (portrait -> landscape -> portrait) and re-check."
            ),
            test_data="N/A",
            expected_result="Option is properly aligned with consistent spacing; no odd offsets in rotation.",
        ),
        # 5) Blurred/low-quality coins icon in Refer & Earn
        TestCase(
            id="HIMA-BUGFIX-009",
            title="Refer & Earn: coins icon renders sharp (no blur/pixelation)",
            module="Refer & Earn",
            severity="Low/Medium",
            priority="Low/Medium",
            preconditions="Login and open Refer & Earn screen.",
            steps=(
                "1. Observe the coins icon at normal zoom.\n"
                "2. Screenshot the screen.\n"
                "3. View screenshot and check icon clarity/edges."
            ),
            test_data="N/A",
            expected_result="Coins icon is crisp and clear; no visible blur or low-resolution scaling artifacts.",
        ),
        # 6) Duplicate username in Favorite tab
        TestCase(
            id="HIMA-BUGFIX-010",
            title="Favorites: username displayed only once per favorite item",
            module="Favorites",
            severity="Medium",
            priority="Medium",
            preconditions="Login with a user having at least 3 favorites. Navigate to Favorite tab.",
            steps=(
                "1. Observe each favorite profile card/list item.\n"
                "2. Confirm username appears only once.\n"
                "3. Scroll and verify for additional items."
            ),
            test_data="User with multiple favorite profiles.",
            expected_result="Username is not duplicated anywhere in the favorites list.",
        ),
        # 7) Help & Support category showing code/technical text
        TestCase(
            id="HIMA-BUGFIX-011",
            title="Help & Support: categories show only user-facing formatted content",
            module="Help & Support",
            severity="Medium",
            priority="Medium",
            preconditions="Login and open Help & Support. Have at least 2 categories available.",
            steps=(
                "1. Open first category.\n"
                "2. Scan content for code/markup/technical strings.\n"
                "3. Open second category and repeat."
            ),
            test_data="Any categories present in Help & Support.",
            expected_result="Only user-readable help content is displayed; no code blocks, HTML tags, or technical text.",
        ),
        # 8) Incomplete help text above Delete Account button (“Need help? please write to”)
        TestCase(
            id="HIMA-BUGFIX-012",
            title="Delete Account screen: help text is complete and correctly capitalized",
            module="Profile > Account Settings > Delete Account",
            severity="Low",
            priority="Low",
            preconditions="Login and navigate to Profile > Account Settings > Delete Account.",
            steps="1. Read the help text shown above the Delete Account button.",
            test_data="N/A",
            expected_result=(
                "Help text is complete, grammatically correct, and includes a clear contact method (e.g., support email)."
            ),
        ),
        # 9) Delete Account button theme mismatch (text/background)
        TestCase(
            id="HIMA-BUGFIX-013",
            title="Delete Account button follows app theme colors",
            module="Profile > Account Settings > Delete Account",
            severity="Low",
            priority="Low",
            preconditions="Login and navigate to Delete Account screen.",
            steps=(
                "1. Observe Delete Account button background color.\n"
                "2. Observe Delete Account button text color.\n"
                "3. Compare with other themed primary/destructive buttons in the app."
            ),
            test_data="N/A",
            expected_result="Delete Account button colors match the expected theme (e.g., white text and/or pink background).",
        ),
        # 10) Missing exclamation mark warning icon in Delete Account confirmation popup
        TestCase(
            id="HIMA-BUGFIX-014",
            title="Delete Account confirmation popup shows warning icon",
            module="Profile > Account Settings > Delete Account",
            severity="Low",
            priority="Low",
            preconditions="Login and navigate to Delete Account screen.",
            steps=(
                "1. Tap Delete Account.\n"
                "2. Observe the confirmation popup.\n"
                "3. Verify warning/exclamation icon is present."
            ),
            test_data="N/A",
            expected_result="Confirmation popup displays the warning icon alongside the warning message.",
        ),
        # 11) Missing exclamation mark icon on Delete Account screen warning message
        TestCase(
            id="HIMA-BUGFIX-015",
            title="Delete Account screen warning message shows exclamation icon",
            module="Profile > Account Settings > Delete Account",
            severity="Low",
            priority="Low",
            preconditions="Login and open Delete Account screen.",
            steps="1. Observe the warning/important information message area on the screen.",
            test_data="N/A",
            expected_result="Warning area includes the exclamation/warning icon displayed alongside the text.",
        ),
        # 12) Account deletion reasons list not scrollable
        TestCase(
            id="HIMA-BUGFIX-016",
            title="Delete Account: reasons list is scrollable to reach all options",
            module="Profile > Account Settings > Delete Account",
            severity="Medium",
            priority="Medium",
            preconditions="Login and open Delete Account screen with enough reasons to exceed viewport.",
            steps=(
                "1. Select all visible reasons.\n"
                "2. Attempt to scroll down.\n"
                "3. Verify additional reasons become visible and selectable."
            ),
            test_data="Delete reasons list long enough to require scroll.",
            expected_result="Screen/container scrolls; user can view and select all reasons.",
        ),
        # 13) Layout breaks when multiple delete reasons selected
        TestCase(
            id="HIMA-BUGFIX-017",
            title="Delete Account: layout remains stable when multiple reasons selected",
            module="Profile > Account Settings > Delete Account",
            severity="Medium",
            priority="Medium",
            preconditions="Login and open Delete Account screen.",
            steps=(
                "1. Select 1 reason and observe layout.\n"
                "2. Select 3+ reasons and observe layout.\n"
                "3. Scroll and confirm layout remains consistent (no overlaps/incorrect spacing)."
            ),
            test_data="N/A",
            expected_result="UI remains properly aligned; no broken layout when selecting multiple reasons.",
        ),
        # 14) Incorrect contact email in Help & Support / Contact Us
        TestCase(
            id="HIMA-BUGFIX-018",
            title="Help & Support > Contact Us shows correct support email address",
            module="Help & Support",
            severity="Medium",
            priority="High",
            preconditions="Login and navigate to Help & Support > Contact Us.",
            steps=(
                "1. Locate displayed support email address.\n"
                "2. Verify it matches the expected/official support email configured for the app."
            ),
            test_data="Expected support email (from product/ops).",
            expected_result="Displayed email address is correct and valid.",
            notes="If your team has a canonical support email, verify exact match.",
        ),
        # 15) Logout requires multiple clicks
        TestCase(
            id="HIMA-BUGFIX-019",
            title="Logout triggers on single tap",
            module="Profile / Settings",
            severity="Medium",
            priority="High",
            preconditions="Login and navigate to Profile / Settings where Logout button is present.",
            steps=(
                "1. Tap Logout once.\n"
                "2. Observe whether user is logged out and redirected as expected."
            ),
            test_data="N/A",
            expected_result="User logs out immediately on first tap (no need for multiple taps).",
        ),
        TestCase(
            id="HIMA-BUGFIX-020",
            title="Logout: prevent double-tap issues (no stuck state)",
            module="Profile / Settings",
            severity="Medium",
            priority="High",
            preconditions="Login and open Profile / Settings screen.",
            steps=(
                "1. Tap Logout rapidly 2-3 times.\n"
                "2. Observe app behavior during logout transition."
            ),
            test_data="N/A",
            expected_result="App handles repeated taps gracefully; logout proceeds once without glitches/crash.",
            notes="Regression check for click debouncing/disabled state.",
        ),
        # 16) Verify OTP button enabled even when OTP empty
        TestCase(
            id="HIMA-BUGFIX-021",
            title="OTP: Verify button disabled when OTP field is empty",
            module="Login / OTP",
            severity="Medium",
            priority="High",
            preconditions="On login screen, enter valid phone number and tap Send OTP. Do not enter OTP.",
            steps=(
                "1. Tap Send OTP.\n"
                "2. Do not enter OTP.\n"
                "3. Observe Verify OTP button state."
            ),
            test_data="Valid phone number.",
            expected_result="Verify OTP button remains disabled until OTP input is provided.",
        ),
        TestCase(
            id="HIMA-BUGFIX-022",
            title="OTP: Verify button enables only for valid OTP length",
            module="Login / OTP",
            severity="Medium",
            priority="High",
            preconditions="Send OTP successfully and stay on OTP verification screen.",
            steps=(
                "1. Enter 1-2 digits of OTP.\n"
                "2. Observe Verify OTP.\n"
                "3. Enter full OTP length.\n"
                "4. Observe Verify OTP enables.\n"
                "5. Clear OTP field and observe button disables again."
            ),
            test_data="OTP length per app requirement (e.g., 4/6 digits).",
            expected_result="Verify OTP enables only when OTP meets required length/format; disables when cleared/incomplete.",
        ),
    ]


def write_xlsx(out_path: Path, testcases: list[TestCase]) -> None:
    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception as e:  # pragma: no cover
        raise SystemExit(
            "Missing dependency: openpyxl. Install it with: pip install openpyxl"
        ) from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bug Fix Test Cases"

    headers = [
        "Test Case ID",
        "Title",
        "Module",
        "Severity",
        "Priority",
        "Preconditions",
        "Steps",
        "Test Data",
        "Expected Result",
        "Actual Result",
        "Notes",
    ]

    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3F51B5")  # Indigo-ish
    header_alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    body_alignment = Alignment(vertical="top", wrap_text=True)
    for tc in testcases:
        ws.append(
            [
                tc.id,
                tc.title,
                tc.module,
                tc.severity,
                tc.priority,
                tc.preconditions,
                tc.steps,
                tc.test_data,
                tc.expected_result,
                tc.actual_result,
                tc.notes,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = body_alignment

    # Freeze header row and enable filters
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{ws.max_row}"

    # Set column widths (tuned for readability in Excel)
    widths = {
        "A": 16,  # ID
        "B": 46,  # Title
        "C": 26,  # Module
        "D": 12,  # Severity
        "E": 12,  # Priority
        "F": 38,  # Preconditions
        "G": 44,  # Steps
        "H": 22,  # Test Data
        "I": 46,  # Expected
        "J": 22,  # Actual
        "K": 26,  # Notes
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Make header row taller
    ws.row_dimensions[1].height = 26

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    out_file = repo_root / "qa" / "testcases" / f"HIMA_BugFixes_TestCases_{_today_stamp()}.xlsx"
    testcases = build_testcases()
    write_xlsx(out_file, testcases)
    print(str(out_file))


if __name__ == "__main__":
    main()

